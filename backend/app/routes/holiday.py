from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import Optional
import io, csv
from datetime import date as date_type, datetime as datetime_type
from fastapi.responses import StreamingResponse

from app.database.session import get_db
from app.core.dependencies import get_current_user
from app.schemas.holiday import HolidayCreate, HolidayUpdate, HolidayOut, HolidayBulkDeleteRequest
from app.services import holiday_service
from app.models.holiday import Holiday, HolidayType
from app.services.notification_service import notify_all_employees, ensure_tomorrow_holiday_notifications

router = APIRouter(prefix="/holidays", tags=["Holidays"])


# ─── LIST ──────────────────────────────────────────────────────────────────────
@router.get("/", response_model=list[HolidayOut])
def list_holidays(
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None),
    department: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    return holiday_service.get_all_holidays(
        db,
        year=year,
        month=month,
        department=department,
        holiday_type=type,
    )


# ─── CREATE ────────────────────────────────────────────────────────────────────
@router.post("/", response_model=HolidayOut, status_code=201)
def create_holiday(
    data: HolidayCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    holiday = holiday_service.create_holiday(db, data)
    notify_all_employees(
        db,
        title="New holiday added",
        message=f"Holiday declared on {holiday.date}: {holiday.name}",
        event_type="holiday_added",
        reference_type="holiday",
        reference_id=holiday.id,
        created_by=getattr(current_user, "id", None)
    )
    ensure_tomorrow_holiday_notifications(db)
    return holiday


# ─── GET ONE ───────────────────────────────────────────────────────────────────
@router.get("/{holiday_id}", response_model=HolidayOut)
def get_holiday(
    holiday_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    holiday = holiday_service.get_holiday_by_id(db, holiday_id)
    if not holiday:
        raise HTTPException(status_code=404, detail="Holiday not found")
    return holiday


# ─── UPDATE ────────────────────────────────────────────────────────────────────
@router.put("/{holiday_id}", response_model=HolidayOut)
def update_holiday(
    holiday_id: int,
    data: HolidayUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    holiday = holiday_service.update_holiday(db, holiday_id, data)
    if not holiday:
        raise HTTPException(status_code=404, detail="Holiday not found")
    ensure_tomorrow_holiday_notifications(db)
    return holiday


# ─── DELETE ONE ────────────────────────────────────────────────────────────────
@router.delete("/{holiday_id}", status_code=204)
def delete_holiday(
    holiday_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    success = holiday_service.delete_holiday(db, holiday_id)
    if not success:
        raise HTTPException(status_code=404, detail="Holiday not found")


# ─── BULK DELETE ───────────────────────────────────────────────────────────────
@router.delete("/", status_code=200)
def bulk_delete_holidays(
    payload: HolidayBulkDeleteRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    deleted = holiday_service.bulk_delete_holidays(db, payload.ids)
    return {"deleted": deleted}


# ─── EXPORT CSV ────────────────────────────────────────────────────────────────
@router.get("/export/csv")
def export_holidays_csv(
    year: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    holidays = holiday_service.get_all_holidays(db, year=year)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Holiday Name", "Date", "Type", "Department", "Repeat Yearly"])

    for h in holidays:
        writer.writerow([h.id, h.name, str(h.date), h.type.value, h.department, h.repeat_yearly])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=holidays_{year or 'all'}.csv"},
    )


# ─── CHECK DATE IS HOLIDAY ─────────────────────────────────────────────────────
@router.get("/check/{date_str}")
def check_date_holiday(
    date_str: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Returns list of holidays on a specific date. Used by frontend attendance calendar."""
    from datetime import date as date_type
    try:
        target = date_type.fromisoformat(date_str)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    holidays = holiday_service.get_holidays_for_date(db, target)
    return [HolidayOut.model_validate(h) for h in holidays]


@router.post("/bulk-upload")
def bulk_upload_holidays(
    file: UploadFile = File(...),
    target_year: Optional[int] = Form(default=None),
    target_month: Optional[int] = Form(default=None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")

    if target_month is not None and (target_month < 1 or target_month > 12):
        raise HTTPException(status_code=400, detail="target_month must be between 1 and 12")

    if (target_year is None) != (target_month is None):
        raise HTTPException(status_code=400, detail="Select both target year and target month together")

    file_bytes = file.file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    def normalize_header(value: str) -> str:
        return str(value or "").strip().lower().replace(" ", "_")

    def parse_bool(value) -> bool:
        return str(value or "").strip().lower() in {"true", "1", "yes", "y"}

    def parse_type(value) -> HolidayType:
        raw = str(value or "").strip().lower()
        mapping = {
            "full_day": HolidayType.full_day,
            "full day": HolidayType.full_day,
            "first_half": HolidayType.first_half,
            "first half": HolidayType.first_half,
            "second_half": HolidayType.second_half,
            "second half": HolidayType.second_half,
        }
        return mapping.get(raw, HolidayType.full_day)

    def parse_date_value(raw_value, raw_day: str | None) -> date_type:
        resolved_date = None
        if isinstance(raw_value, datetime_type):
            resolved_date = raw_value.date()
        elif isinstance(raw_value, date_type):
            resolved_date = raw_value

        if target_year is not None and target_month is not None:
            day_number = None
            if raw_day:
                try:
                    day_number = int(str(raw_day).strip())
                except Exception:
                    day_number = None

            if day_number is None:
                if resolved_date is not None:
                    day_number = resolved_date.day
                elif not raw_value:
                    raise ValueError("Either date or day is required")
                else:
                    parsed = date_type.fromisoformat(str(raw_value).strip())
                    day_number = parsed.day

            return date_type(target_year, target_month, day_number)

        if resolved_date is not None:
            return resolved_date

        if not raw_value:
            raise ValueError("Date is required")
        return date_type.fromisoformat(str(raw_value).strip())

    rows: list[dict] = []
    if filename.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({normalize_header(k): (v.strip() if isinstance(v, str) else v) for k, v in (row or {}).items()})
    else:
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(status_code=500, detail="Excel upload requires openpyxl package")

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        headers = [normalize_header(h) for h in all_rows[0]]
        for r in all_rows[1:]:
            if not any(cell is not None and str(cell).strip() for cell in r):
                continue
            mapped = {}
            for idx, value in enumerate(r):
                key = headers[idx] if idx < len(headers) else f"col_{idx}"
                mapped[key] = str(value).strip() if value is not None else ""
            rows.append(mapped)

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    created = 0
    updated = 0
    failed = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        try:
            name = row.get("name") or row.get("holiday_name") or ""
            day = row.get("day")
            raw_date = row.get("date")
            holiday_date = parse_date_value(raw_date, day)
            if holiday_date < date_type.today():
                raise ValueError("Past dates are not allowed")

            holiday_type = parse_type(row.get("type"))
            department = row.get("department") or "All"
            repeat_yearly = parse_bool(row.get("repeat_yearly"))

            if not name:
                raise ValueError("Holiday name is required")

            existing = db.query(Holiday).filter(
                Holiday.name == name,
                Holiday.date == holiday_date,
                Holiday.department == department
            ).first()

            if existing:
                existing.type = holiday_type
                existing.repeat_yearly = repeat_yearly
                holiday_service.update_holiday(
                    db,
                    existing.id,
                    HolidayUpdate(
                        name=existing.name,
                        date=existing.date,
                        type=existing.type,
                        department=existing.department,
                        repeat_yearly=existing.repeat_yearly
                    )
                )
                updated += 1
            else:
                holiday_service.create_holiday(
                    db,
                    HolidayCreate(
                        name=name,
                        date=holiday_date,
                        type=holiday_type,
                        department=department,
                        repeat_yearly=repeat_yearly
                    )
                )
                latest_holiday = db.query(Holiday).filter(
                    Holiday.name == name,
                    Holiday.date == holiday_date,
                    Holiday.department == department
                ).order_by(Holiday.id.desc()).first()
                if latest_holiday:
                    notify_all_employees(
                        db,
                        title="New holiday added",
                        message=f"Holiday declared on {latest_holiday.date}: {latest_holiday.name}",
                        event_type="holiday_added",
                        reference_type="holiday",
                        reference_id=latest_holiday.id,
                        created_by=getattr(current_user, "id", None)
                    )
                created += 1

        except Exception as exc:
            failed += 1
            errors.append(f"Row {idx}: {exc}")

    ensure_tomorrow_holiday_notifications(db)
    return {
        "message": "Bulk upload processed",
        "created": created,
        "updated": updated,
        "failed": failed,
        "errors": errors[:20]
    }
